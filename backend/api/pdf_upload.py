from fastapi import APIRouter, UploadFile, File, HTTPException, Depends
from sqlalchemy.orm import Session
from sqlalchemy import text
import openpyxl
from datetime import datetime, date
import tempfile
import os
import re
from models.database import get_db
from models.models import Warehouse, Product, Inventory

router = APIRouter()

# ── HELPER FUNCTIONS ──────────────────────────────────────────

COLUMN_ALIASES = {
    "product id": "product_id",
    "product_id": "product_id",
    "sku": "product_id",
    "product name": "product_name",
    "product_name": "product_name",
    "name": "product_name",
    "category": "category",
    "stock": "stock",
    "current stock": "stock",
    "available quantity": "stock",
    "units to produce": "units_to_produce",
    "restock date": "restock_date",
    "expected restock date": "restock_date",
    "dispatch limit": "dispatch_limit",
    "dispatch limit/day": "dispatch_limit",
    "dispatch limit per day": "dispatch_limit",
    "warehouse": "warehouse",
}


def clean_product_id(value):
    """Ensures 1001 doesn't parse as 1001.0"""
    if value is None:
        return ""
    val_str = str(value).strip()
    if val_str.endswith(".0"):
        return val_str[:-2]
    return val_str


def parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text_val = str(value).strip()
    match = re.search(r"\d{4}-\d{2}-\d{2}", text_val)
    if match:
        try:
            return datetime.strptime(match.group(), "%Y-%m-%d").date()
        except ValueError:
            return None
    match = re.search(r"\d{2}/\d{2}/\d{4}", text_val)
    if match:
        try:
            return datetime.strptime(match.group(), "%d/%m/%Y").date()
        except ValueError:
            return None
    return None


def parse_number(value, default=0):
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return int(value)
    match = re.search(r"\d+", str(value))
    return int(match.group()) if match else default


def extract_warehouse_name(sheet_name):
    match = re.search(r"warehouse\s*\d+", sheet_name, re.IGNORECASE)
    if match:
        name = re.sub(r"\s+", " ", match.group()).strip()
        return name[0].upper() + name[1:]
    return None


def build_header_map(header_row):
    header_map = {}
    has_explicit_product_id = False
    has_product_name = False
    
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        text_val = str(cell).strip().lower()
        text_clean = re.sub(r"[\s_-]+", " ", text_val)
        
        if text_clean == "id":
            continue
            
        if text_clean in ["product id", "product_id", "sku"]:
            header_map[idx] = "product_id"
            has_explicit_product_id = True
        elif text_clean in ["product name", "product_name", "name"]:
            header_map[idx] = "product_name"
            has_product_name = True
        elif text_clean in ["current stock", "available quantity", "stock"]:
            header_map[idx] = "stock"
        elif text_clean in ["category"]:
            header_map[idx] = "category"
        elif text_clean in ["units to produce"]:
            header_map[idx] = "units_to_produce"
        elif text_clean in ["restock date", "expected restock date"]:
            header_map[idx] = "restock_date"
        elif text_clean in ["dispatch limit", "dispatch limit/day", "dispatch limit per day"]:
            header_map[idx] = "dispatch_limit"
        elif text_clean in ["warehouse"]:
            header_map[idx] = "warehouse"

    if has_explicit_product_id and has_product_name:
        return header_map
    return {}


def row_to_record(row, header_map):
    record = {}
    for idx, key in header_map.items():
        if idx < len(row):
            record[key] = row[idx]
    return record


def ensure_is_latest_column(db):
    try:
        db.execute(text("ALTER TABLE warehouses ADD COLUMN is_latest INTEGER DEFAULT 0"))
        db.commit()
    except Exception:
        pass


# ── MAIN EXCEL UPLOAD ENDPOINT ─────────────────────────────────

@router.post("/upload-excel")
async def upload_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    if not (file.filename.endswith(".xlsx") or file.filename.endswith(".xlsm")):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xlsm files allowed!")

    ensure_is_latest_column(db)

    # 🔄 AUTOMATIC CLEAR ON NEW UPLOAD:
    print("🧹 New Excel detected. Clearing old inventory metrics while preserving users...")
    try:
        db.query(Inventory).delete()
        db.query(Product).delete()
        db.query(Warehouse).delete()
        db.commit()
    except Exception as clear_err:
        print(f"⚠️ Notice during clear loop: {clear_err}")
        db.rollback()

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name

    try:
        warehouses_added = 0
        warehouses_updated = 0
        products_added = 0
        inventory_added = 0
        inventory_updated = 0
        errors = []
        extracted_data = []
        latest_warehouse_ids = []

        try:
            workbook = openpyxl.load_workbook(tmp_path, data_only=True)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Could not read Excel file: {str(e)}")

        warehouse_cache = {}

        def get_or_create_warehouse(wh_name):
            nonlocal warehouses_added, warehouses_updated
            if wh_name in warehouse_cache:
                return warehouse_cache[wh_name]
            existing = db.query(Warehouse).filter(Warehouse.name == wh_name).first()
            if not existing:
                wh = Warehouse(name=wh_name, location=wh_name)
                db.add(wh)
                db.commit()
                db.refresh(wh)
                warehouse_cache[wh_name] = wh.id
                warehouses_added += 1
                return wh.id
            else:
                warehouse_cache[wh_name] = existing.id
                warehouses_updated += 1
                return existing.id

        any_rows_processed = False

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue

            header_row_idx = None
            header_map = {}
            for r_idx, row in enumerate(rows):
                candidate_map = build_header_map(row)
                if len(candidate_map) >= 2:
                    header_row_idx = r_idx
                    header_map = candidate_map
                    break

            if header_row_idx is None:
                continue

            sheet_warehouse_name = extract_warehouse_name(sheet_name)

            for row in rows[header_row_idx + 1:]:
                if row is None or all(c is None for c in row):
                    continue

                record = row_to_record(row, header_map)
                wh_name = sheet_warehouse_name
                if not wh_name and record.get("warehouse"):
                    raw_wh = str(record["warehouse"]).strip()
                    wh_name = extract_warehouse_name(raw_wh) or raw_wh

                if not wh_name:
                    continue

                product_id_raw = clean_product_id(record.get("product_id"))
                product_name = str(record.get("product_name")).strip() if record.get("product_name") is not None else ""

                if not product_name or not product_id_raw:
                    continue

                category = str(record.get("category")).strip() if record.get("category") else "General"
                current_stock = parse_number(record.get("stock"), default=0)
                units_to_produce = parse_number(record.get("units_to_produce"), default=0)
                restock_date = parse_date(record.get("restock_date"))
                dispatch_limit = parse_number(record.get("dispatch_limit"), default=50)

                try:
                    warehouse_id = get_or_create_warehouse(wh_name)

                    if warehouse_id not in latest_warehouse_ids:
                        latest_warehouse_ids.append(warehouse_id)

                    existing_product = db.query(Product).filter(Product.name == product_name).first()
                    if not existing_product:
                        product = Product(name=product_name, description=category, unit_price=0.00)
                        db.add(product)
                        db.commit()
                        db.refresh(product)
                        actual_product_id = product.id
                        products_added += 1
                    else:
                        existing_product.description = category
                        db.commit()
                        actual_product_id = existing_product.id

                    inv = Inventory(
                        warehouse_id=warehouse_id,
                        product_id=actual_product_id,
                        pdf_product_id=product_id_raw,  
                        available_quantity=current_stock,
                        dispatch_limit=dispatch_limit,
                        dispatched_today=0,
                        units_to_produce=units_to_produce,
                        restock_date=restock_date
                    )
                    db.add(inv)
                    db.commit()
                    db.refresh(inv)
                    inventory_added += 1

                    any_rows_processed = True

                except Exception as row_error:
                    errors.append(str(row_error))
                    continue

        db.execute(text("UPDATE warehouses SET is_latest = 0"))
        for wid in latest_warehouse_ids:
            db.execute(text(f"UPDATE warehouses SET is_latest = 1 WHERE id = {wid}"))
        db.commit()

        # 🔄 VIP BACKLOG CLEARANCE PLUGGED HERE
        try:
            from api.routes import clear_vip_backlog_after_upload
            print("🔄 Processing rolling VIP backlog subtraction limits...")
            clear_vip_backlog_after_upload(db)
            print("✅ Backlog rolling queue cleared safely!")
        except Exception as e:
            print(f"⚠️ Backlog clearance execution failed: {e}")

        return {"message": "Success"}

    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


# ── GET INVENTORY ENDPOINT (FIXED ROUTING DISPLAY VALUE) ───────

@router.get("/inventory/all")
def get_all_inventory(db: Session = Depends(get_db)):
    ensure_is_latest_column(db)
    
    warehouse_objs = db.query(Warehouse).all()
    result = []
    
    for wh in warehouse_objs:
        inventory = db.query(Inventory).filter(Inventory.warehouse_id == wh.id).all()
        products = []
        for inv in inventory:
            product = db.query(Product).filter(Product.id == inv.product_id).first()
            if product:
                products.append({
                    "inventory_id": inv.id,
                    "product_id": inv.pdf_product_id if inv.pdf_product_id else str(product.id),
                    "db_product_id": product.id,
                    "product_name": product.name,
                    "category": product.description,
                    "available_quantity": inv.available_quantity,
                    "units_to_produce": inv.units_to_produce,
                    "dispatch_limit": inv.dispatch_limit,
                    "dispatched_today": inv.dispatched_today,
                    "restock_date": str(inv.restock_date) if inv.restock_date else None,
                    "warehouse_id": inv.warehouse_id
                })
        result.append({
            "warehouse_id": wh.id,
            "warehouse_name": wh.name,
            "location": wh.location,
            "total_products": len(products),
            "products": products
        })
    return result


# ── CLEAR ALL DATA ENDPOINT ────────────────────────────────────

@router.delete("/clear/all")
def clear_all_data(db: Session = Depends(get_db)):
    db.query(Inventory).delete()
    db.query(Product).delete()
    db.query(Warehouse).delete()
    db.commit()
    return {"message": "✅ Database wiped successfully! Ready for clean upload."}


@router.post("/inventory/upload")
async def upload_inventory_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    return await upload_excel(file=file, db=db)